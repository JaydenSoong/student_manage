import openpyxl

'''
用 openpyxl 读取 excel 文件的操作
'''
class ReadExcel:
    def __init__(self, file_path):
        # 1.加载 excel 文件
        self.workbook = openpyxl.load_workbook(file_path)
        # 2.获取当前工作薄
        self.worksheet = self.workbook.active

    # 读取的方法
    def get_data(self):
        # 3.逐行读取
        # 将最后读取的数据以列表的方式保存起来
        data = []
        for row in self.worksheet.iter_rows():
            row_data = []
            # 遍历行数据
            for cell in row:
                row_data.append(cell.value)
            data.append(row_data)
        return  data

'''
用 openpyxl 写入 excel 文件的操作
'''
class WriteExcel:
    def __init__(self, file_path, data):
        # 获取文件路径
        self.file_path = file_path
        # 获取数据
        self.data = data
        # 创建一个对象用来创建一个工作表
        self.workbook = openpyxl.Workbook()
        # 获取当前工作薄
        self.worksheet = self.workbook.active

    def write_data(self):

        # 逐行写入
        for row in self.data:
            self.worksheet.append(row)

        self.workbook.save(self.file_path)